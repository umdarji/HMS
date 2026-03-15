from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, Count, Q
from .models import Patient, Doctor, Staff, OPDAppointment, IPDAdmission, Payment, Bed
import json

@login_required
def dashboard_stats_api(request):
    """
    API endpoint for real-time admin dashboard statistics
    Returns JSON data for AJAX consumption
    """
    if request.user.user_type != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    today = timezone.localdate()
    
    # Basic counts
    total_patients = Patient.objects.count()
    total_doctors = Doctor.objects.count()
    total_staff = Staff.objects.count()
    total_appointments = OPDAppointment.objects.count()
    
    # Today's appointments
    today_appointments = OPDAppointment.objects.filter(
        appointment_date__date=today
    ).count()
    
    # OPD vs IPD counts
    opd_count = OPDAppointment.objects.count()
    ipd_count = IPDAdmission.objects.filter(status='Admitted').count()
    
    # Bed statistics
    total_beds = Bed.objects.count()
    available_beds = Bed.objects.filter(status='Available').count()
    occupied_beds = Bed.objects.filter(status='Occupied').count()
    
    # Revenue calculation
    payment_revenue = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    opd_revenue = OPDAppointment.objects.aggregate(total=Sum('fee'))['total'] or 0
    total_revenue = float(payment_revenue) + float(opd_revenue)
    
    # Today's revenue
    today_payment_revenue = Payment.objects.filter(
        payment_date__date=today
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    today_opd_revenue = OPDAppointment.objects.filter(
        appointment_date__date=today
    ).aggregate(total=Sum('fee'))['total'] or 0
    
    today_revenue = float(today_payment_revenue) + float(today_opd_revenue)
    
    # Recent appointments (last 10)
    recent_appointments = OPDAppointment.objects.select_related(
        'patient', 'doctor'
    ).order_by('-appointment_date')[:10]
    
    appointments_list = []
    for appointment in recent_appointments:
        appointments_list.append({
            'patient_name': appointment.patient.name,
            'doctor_name': appointment.doctor.name if appointment.doctor else 'N/A',
            'date': appointment.appointment_date.strftime('%b %d, %Y'),
            'time': appointment.appointment_date.strftime('%I:%M %p'),
            'status': appointment.status,
            'status_badge': get_status_badge(appointment.status)
        })
    
    # Doctor availability stats
    available_doctors = Doctor.objects.filter(availability_status='Available').count()
    
    # Response data
    data = {
        'counts': {
            'total_patients': total_patients,
            'total_doctors': total_doctors,
            'total_staff': total_staff,
            'total_appointments': total_appointments,
            'today_appointments': today_appointments,
            'opd_count': opd_count,
            'ipd_count': ipd_count,
            'available_doctors': available_doctors,
        },
        'beds': {
            'total': total_beds,
            'available': available_beds,
            'occupied': occupied_beds,
        },
        'revenue': {
            'total': round(total_revenue, 2),
            'today': round(today_revenue, 2),
        },
        'recent_appointments': appointments_list,
        'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    
    return JsonResponse(data)


def get_status_badge(status):
    """Helper function to return Bootstrap badge class for appointment status"""
    status_map = {
        'Pending': 'warning',
        'Completed': 'success',
        'Cancelled': 'danger',
        'Confirmed': 'info',
    }
    return status_map.get(status, 'secondary')


@login_required
def reports_stats_api(request):
    """
    API endpoint for real-time reports & analytics page statistics
    Returns JSON data for AJAX consumption
    """
    if request.user.user_type != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    today = timezone.localdate()
    
    # Calculate total revenue
    payment_revenue = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    opd_revenue = OPDAppointment.objects.aggregate(total=Sum('fee'))['total'] or 0
    total_revenue = float(payment_revenue) + float(opd_revenue)
    
    # Total patients count
    total_patients = Patient.objects.count()
    
    # Total appointments count
    total_appointments = OPDAppointment.objects.count()
    
    # Calculate satisfaction rate (based on completed appointments)
    total_appts = OPDAppointment.objects.count()
    completed_appts = OPDAppointment.objects.filter(status='Completed').count()
    satisfaction_rate = round((completed_appts / total_appts * 100) if total_appts > 0 else 0, 1)
    
    # Department-wise statistics
    from django.db.models import Count
    department_stats = Department.objects.annotate(
        appointment_count=Count('doctor__opdappointment')
    ).values('name', 'appointment_count').order_by('-appointment_count')[:5]
    
    # Monthly revenue trend (last 6 months)
    from datetime import timedelta
    from django.db.models.functions import TruncMonth
    
    six_months_ago = today - timedelta(days=180)
    monthly_revenue = Payment.objects.filter(
        payment_date__date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('payment_date')
    ).values('month').annotate(
        revenue=Sum('amount')
    ).order_by('month')
    
    monthly_data = []
    for item in monthly_revenue:
        monthly_data.append({
            'month': item['month'].strftime('%b %Y'),
            'revenue': float(item['revenue'] or 0)
        })
    
    # Response data
    data = {
        'stats': {
            'total_revenue': round(total_revenue, 2),
            'total_patients': total_patients,
            'total_appointments': total_appointments,
            'satisfaction_rate': satisfaction_rate,
        },
        'department_stats': list(department_stats),
        'monthly_revenue': monthly_data,
        'timestamp': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    
    return JsonResponse(data)


@login_required
def financial_report_pdf(request):
    """Generate PDF financial report"""
    if request.user.user_type != 'admin':
        return redirect('login')
    
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=20, 
                                 textColor=colors.HexColor('#2c3e50'), spaceAfter=20, 
                                 alignment=TA_CENTER, fontName='Helvetica-Bold')
    
    title = Paragraph("Financial Report - Hospital Management System", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Revenue Summary
    payment_revenue = Payment.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    opd_revenue = OPDAppointment.objects.aggregate(Sum('fee'))['fee__sum'] or 0
    total_revenue = float(payment_revenue) + float(opd_revenue)
    
    summary_data = [
        ['Revenue Source', 'Amount (₹)'],
        ['OPD Fees', f'{opd_revenue:.2f}'],
        ['Other Payments', f'{payment_revenue:.2f}'],
        ['Total Revenue', f'{total_revenue:.2f}']
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8f4f8')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Recent Payments
    subtitle = Paragraph("Recent Transactions", styles['Heading2'])
    elements.append(subtitle)
    elements.append(Spacer(1, 10))
    
    payments = Payment.objects.all().order_by('-payment_date')[:10]
    payment_data = [['Date', 'Patient', 'Amount (₹)', 'Method']]
    
    for payment in payments:
        payment_data.append([
            payment.payment_date.strftime('%Y-%m-%d'),
            payment.patient.name[:20],
            f'{payment.amount:.2f}',
            payment.payment_method
        ])
    
    payment_table = Table(payment_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch])
    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ecf0f1')]),
    ]))
    
    elements.append(payment_table)
    
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=10, 
                                  textColor=colors.grey, alignment=TA_CENTER)
    footer = Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style)
    elements.append(Spacer(1, 20))
    elements.append(footer)
    
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="financial_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    response.write(pdf)
    
    return response


@login_required
def patient_records_excel(request):
    """Generate Excel file with patient records"""
    if request.user.user_type != 'admin':
        return redirect('login')
    
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Records"
    
    # Headers
    headers = ['Patient ID', 'Name', 'Age', 'Gender', 'Blood Group', 'Phone', 'Email', 'Registration Date']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add patient data
    patients = Patient.objects.all().order_by('-created_at')
    for patient in patients:
        ws.append([
            patient.patient_id,
            patient.name,
            patient.age,
            patient.gender,
            patient.blood_group or 'N/A',
            patient.phone,
            patient.email,
            patient.created_at.strftime('%Y-%m-%d')
        ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="patient_records_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


@login_required
def staff_performance_csv(request):
    """Generate CSV file with staff performance data"""
    if request.user.user_type != 'admin':
        return redirect('login')
    
    from django.http import HttpResponse
    import csv
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="staff_performance_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Staff ID', 'Name', 'Role', 'Department', 'Joining Date', 'Email', 'Phone'])
    
    staff_members = Staff.objects.all().order_by('-joining_date')
    for staff in staff_members:
        writer.writerow([
            staff.staff_id,
            staff.name,
            staff.role,
            staff.department.name if staff.department else 'N/A',
            staff.joining_date.strftime('%Y-%m-%d'),
            staff.email,
            staff.phone
        ])
    
    return response

